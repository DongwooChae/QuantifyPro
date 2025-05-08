import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# 1) 엑셀 파일 및 시트 로드(매크로 유지)
file_path = 'company_code.xlsx'
wb = load_workbook(filename=file_path)
ws = wb.active




code_xlsx = pd.read_excel('company_code.xlsx', sheet_name=0, index_col = 'symbol')
symbols = code_xlsx.index.tolist()


# url = f'https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode={code}&cID=AA&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=701'
url = f'https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A005930&cID=AA&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=701'
컴퍼니가이드 = requests.get('https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode=A005930&cID=AA&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=701')
컴퍼니가이드 = requests.get(f'https://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?pGB=1&gicode={symbol}&cID=AA&MenuYn=Y&ReportGB=&NewMenuID=11&stkGb=701')

# print(컴퍼니가이드.content)
# print(컴퍼니가이드.status_code)

soup = BeautifulSoup(컴퍼니가이드.content, 'html.parser')

# print(soup.find_all('h3', id='bizSummaryHeader')[0].text)
title = (soup.find_all('h3', id='bizSummaryHeader')[0].text)
# print(soup.find_all('ul', id='bizSummaryContent')[0].text)
content = (soup.find_all('ul', id='bizSummaryContent')[0])
full_content = content.get_text(separator='\n\n', strip=True)

ws['C2'] = f"{title}\n\n{full_content}"

# 셀 안에서 줄바꿈 보이도록 설정
ws['C2'].alignment = Alignment(wrap_text=True)

wb.save(file_path)